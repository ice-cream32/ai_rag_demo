"""从 xlsx 导入手工规则并转换为内部对象。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from app.agent.rule_learning.schemas import ImportReport, ImportStats, ImportWarning, RuleRecord
from app.agent.rule_learning.utils import build_rule_id, norm_header, parse_loose_mapping, pick_first, safe_str

HEADER_ALIASES = {
    "rule_id": {"ruleid", "规则id", "规则编号", "id"},
    "brand": {"brand", "品牌", "品牌名称", "品牌代码"},
    "category": {"category", "类别", "产品类型", "type"},
    "pattern": {"pattern", "regex", "规则", "匹配规则", "编码规则", "料号模式"},
    "field_mapping": {"fieldmapping", "字段映射", "字段map", "mapping"},
    "segment_rules": {"segmentrules", "编码段规则", "段规则", "segments"},
    "enum_values": {"enumvalues", "枚举值", "枚举说明", "枚举", "enums"},
    "remarks": {"remarks", "备注", "说明", "notes"},
}


def _canonical_header(header: str) -> str:
    h = norm_header(header)
    for canonical, aliases in HEADER_ALIASES.items():
        if h in aliases:
            return canonical
    return h


class ExcelRuleImporter:
    """读取 xlsx 并转换为 RuleRecord 列表。"""

    def import_file(self, xlsx_path: str) -> Tuple[List[RuleRecord], ImportReport]:
        path = Path(xlsx_path)
        if not path.exists():
            raise FileNotFoundError(f"xlsx 文件不存在: {xlsx_path}")

        wb = load_workbook(filename=str(path), data_only=True)
        imported_at = datetime.utcnow()
        stats = ImportStats()
        warnings: List[ImportWarning] = []
        rules: List[RuleRecord] = []

        for ws in wb.worksheets:
            if ws.max_row <= 1:
                continue

            raw_headers = [safe_str(c.value) for c in ws[1]]
            canonical_headers = [_canonical_header(h) for h in raw_headers]
            index_map: Dict[str, int] = {
                name: idx for idx, name in enumerate(canonical_headers) if name
            }

            # 跳过缺少核心字段的 sheet
            if "brand" not in index_map and "pattern" not in index_map and "field_mapping" not in index_map:
                warnings.append(ImportWarning(sheet=ws.title, row=1, message="跳过 sheet：未识别规则表头"))
                continue

            for row_idx in range(2, ws.max_row + 1):
                stats.total_rows += 1
                row_values = [ws.cell(row=row_idx, column=i + 1).value for i in range(len(raw_headers))]
                row_dict = {
                    canonical_headers[i]: row_values[i]
                    for i in range(len(canonical_headers))
                    if i < len(row_values) and canonical_headers[i]
                }

                brand = safe_str(pick_first(row_dict, ["brand"], ""))
                category = safe_str(pick_first(row_dict, ["category"], ""))
                pattern = safe_str(pick_first(row_dict, ["pattern"], ""))
                remarks = safe_str(pick_first(row_dict, ["remarks"], ""))

                field_mapping = parse_loose_mapping(pick_first(row_dict, ["field_mapping"], {}))
                segment_rules = parse_loose_mapping(pick_first(row_dict, ["segment_rules"], {}))
                enum_values = parse_loose_mapping(pick_first(row_dict, ["enum_values"], {}))

                # 忽略空行
                if not any([brand, category, pattern, remarks, field_mapping, segment_rules, enum_values]):
                    stats.skipped_rows += 1
                    continue

                rule_id = safe_str(pick_first(row_dict, ["rule_id"], ""))
                if not rule_id:
                    rule_id = build_rule_id(brand=brand, category=category, pattern=pattern, sheet=ws.title, row=row_idx)

                if not brand and not category and not pattern:
                    warnings.append(ImportWarning(sheet=ws.title, row=row_idx, message="缺少 brand/category/pattern，已跳过"))
                    stats.skipped_rows += 1
                    continue

                stats.parsed_rows += 1
                stats.by_brand[brand or "UNKNOWN"] = stats.by_brand.get(brand or "UNKNOWN", 0) + 1

                rules.append(
                    RuleRecord(
                        rule_id=rule_id,
                        brand=brand,
                        category=category,
                        pattern=pattern,
                        field_mapping=field_mapping,
                        segment_rules=segment_rules,
                        enum_values=enum_values,
                        remarks=remarks,
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        imported_at=imported_at,
                    )
                )

        report = ImportReport(
            source_file=path.name,
            imported_at=imported_at,
            sheets=[ws.title for ws in wb.worksheets],
            stats=stats,
            warnings=warnings,
        )
        return rules, report

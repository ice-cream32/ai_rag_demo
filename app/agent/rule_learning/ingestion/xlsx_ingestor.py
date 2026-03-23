"""XLSX ingestor with flexible header semantics (Phase 1)."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from app.agent.rule_learning.utils import norm_header, safe_str


HEADER_SYNONYMS = {
    "brand": {"品牌", "品牌名", "品牌中文名称", "品牌英文名称", "manufacturer", "brand"},
    "category": {"类别", "分类", "类型", "产品类型", "category", "type"},
    "pattern": {"匹配规则", "规则", "regex", "pattern", "命名规则", "料号规则"},
    "capacity_rule": {"容量规则", "容量映射", "容量字符串", "density", "capacity"},
    "suffix_rule": {"后缀规则", "后缀说明", "suffix", "suffixrule"},
    "rule_note": {"规则说明", "描述", "备注", "说明", "rulenotes", "note", "notes"},
}


def _semantic_of(header: str) -> str:
    h = norm_header(header)
    for key, aliases in HEADER_SYNONYMS.items():
        if h in {norm_header(x) for x in aliases}:
            return key
    return "raw"


class XlsxIngestor:
    """Ingest xlsx sheets into unified rule-learning chunks."""

    def ingest(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"xlsx 文件不存在: {file_path}")

        wb = load_workbook(str(path), data_only=True)
        extracted_at = datetime.utcnow().isoformat()
        chunks: List[Dict[str, Any]] = []
        warnings: List[str] = []
        chunk_idx = 0

        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue

            headers = [safe_str(c.value) for c in ws[1]]
            semantics = [_semantic_of(h) if h else "raw" for h in headers]

            for row_idx in range(2, ws.max_row + 1):
                values = [ws.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
                if not any(v not in (None, "") for v in values):
                    continue

                mapped: Dict[str, str] = {}
                raw_map: Dict[str, str] = {}
                for i, value in enumerate(values):
                    val = safe_str(value)
                    if not val:
                        continue
                    header = headers[i] if i < len(headers) else f"col_{i + 1}"
                    sem = semantics[i] if i < len(semantics) else "raw"
                    raw_map[header or f"col_{i + 1}"] = val
                    if sem != "raw" and sem not in mapped:
                        mapped[sem] = val

                if not raw_map:
                    continue

                key_pairs = [f"{k}: {v}" for k, v in mapped.items()] or [f"{k}: {v}" for k, v in raw_map.items()]
                raw_text = "\n".join(key_pairs)

                chunks.append(
                    {
                        "chunk_id": f"xlsx::{path.name}::{ws.title}::{row_idx}::{chunk_idx}",
                        "input_type": "xlsx",
                        "file_name": path.name,
                        "sheet_name": ws.title,
                        "page": None,
                        "section_title": ws.title,
                        "raw_text": raw_text,
                        "metadata": {
                            "row_index": row_idx,
                            "semantic_fields": mapped,
                            "raw_fields": raw_map,
                        },
                        "extracted_at": extracted_at,
                    }
                )
                chunk_idx += 1

            if ws.max_row > 1 and not any(c.get("sheet_name") == ws.title for c in chunks):
                warnings.append(f"sheet={ws.title}: 未提取到有效数据行")

        if not chunks:
            warnings.append("xlsx 未提取到任何 chunk")

        return chunks, warnings

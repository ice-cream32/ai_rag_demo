"""Route multi-format sources to corresponding ingestors (Phase 1)."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import fitz
from openpyxl import load_workbook

from app.agent.rule_learning.utils import norm_header, safe_str


KEYWORDS = ["规则", "映射", "后缀", "前缀", "字段", "示例", "pattern", "regex", "rule"]

HEADER_SYNONYMS = {
    "brand": {"品牌", "品牌名", "品牌中文名称", "品牌英文名称", "manufacturer", "brand"},
    "category": {"类别", "分类", "类型", "产品类型", "category", "type"},
    "pattern": {"匹配规则", "规则", "regex", "pattern", "命名规则", "料号规则"},
    "capacity_rule": {"容量规则", "容量映射", "容量字符串", "density", "capacity"},
    "suffix_rule": {"后缀规则", "后缀说明", "suffix", "suffixrule"},
    "rule_note": {"规则说明", "描述", "备注", "说明", "rulenotes", "note", "notes"},
}


def _split_text_blocks(text: str) -> List[str]:
    blocks = re.split(r"\n\s*\n+", text)
    out: List[str] = []
    for b in blocks:
        t = re.sub(r"\s+", " ", b).strip()
        if t:
            out.append(t)
    return out


def _split_paragraphs(text: str) -> List[str]:
    blocks = re.split(r"\n\s*\n+", text)
    chunks: List[str] = []
    for block in blocks:
        item = re.sub(r"\s+", " ", block).strip()
        if item:
            chunks.append(item)
    return chunks


def _semantic_of(header: str) -> str:
    h = norm_header(header)
    for key, aliases in HEADER_SYNONYMS.items():
        if h in {norm_header(x) for x in aliases}:
            return key
    return "raw"


class TextIngestor:
    """Ingest txt/md/raw text into unified rule-learning chunks."""

    def ingest(self, text: str, file_name: Optional[str] = None, input_type: str = "text") -> Tuple[List[Dict[str, Any]], List[str]]:
        extracted_at = datetime.utcnow().isoformat()
        chunks: List[Dict[str, Any]] = []
        warnings: List[str] = []

        content = (text or "").strip()
        if not content:
            return [], ["文本内容为空"]

        blocks = _split_text_blocks(content)
        current_section = "Text"
        chunk_idx = 0

        for block in blocks:
            # Markdown heading as section title.
            if block.startswith("#"):
                current_section = re.sub(r"^#+\s*", "", block).strip() or current_section

            quality = "rule_like" if any(k.lower() in block.lower() for k in KEYWORDS) else "plain"
            chunks.append(
                {
                    "chunk_id": f"{input_type}::{file_name or 'inline'}::{chunk_idx}",
                    "input_type": input_type,
                    "file_name": file_name or "inline_text",
                    "sheet_name": None,
                    "page": None,
                    "section_title": current_section,
                    "raw_text": block,
                    "metadata": {
                        "quality": quality,
                        "length": len(block),
                    },
                    "extracted_at": extracted_at,
                }
            )
            chunk_idx += 1

        if not chunks:
            warnings.append("文本未提取到任何 chunk")

        return chunks, warnings

    def ingest_file(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文本文件不存在: {file_path}")
        suffix = path.suffix.lower()
        input_type = "md" if suffix == ".md" else "text"
        text = path.read_text(encoding="utf-8", errors="ignore")
        return self.ingest(text=text, file_name=path.name, input_type=input_type)


class PdfIngestor:
    """Ingest pdf pages into unified rule-learning chunks."""

    def ingest(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"pdf 文件不存在: {file_path}")

        extracted_at = datetime.utcnow().isoformat()
        chunks: List[Dict[str, Any]] = []
        warnings: List[str] = []
        chunk_idx = 0

        doc = fitz.open(str(path))
        try:
            for i, page in enumerate(doc):
                page_num = i + 1
                text = page.get_text("text") or ""
                paragraphs = _split_paragraphs(text)

                if not paragraphs:
                    warnings.append(f"page={page_num}: 无可提取文本")
                    continue

                # 普通段落
                for para in paragraphs:
                    chunks.append(
                        {
                            "chunk_id": f"pdf::{path.name}::p{page_num}::{chunk_idx}",
                            "input_type": "pdf",
                            "file_name": path.name,
                            "sheet_name": None,
                            "page": page_num,
                            "section_title": f"Page {page_num}",
                            "raw_text": para,
                            "metadata": {
                                "parser": "fitz.text",
                                "page_index": page_num,
                            },
                            "extracted_at": extracted_at,
                        }
                    )
                    chunk_idx += 1

                # 简易表格线索（包含 | 或连续多空格）
                table_like = [ln.strip() for ln in text.splitlines() if ("|" in ln or re.search(r"\S\s{2,}\S", ln))]
                if table_like:
                    chunks.append(
                        {
                            "chunk_id": f"pdf::{path.name}::p{page_num}::table::{chunk_idx}",
                            "input_type": "pdf",
                            "file_name": path.name,
                            "sheet_name": None,
                            "page": page_num,
                            "section_title": f"Page {page_num} TableLike",
                            "raw_text": "\n".join(table_like[:120]),
                            "metadata": {
                                "parser": "fitz.table_like",
                                "page_index": page_num,
                                "line_count": len(table_like),
                            },
                            "extracted_at": extracted_at,
                        }
                    )
                    chunk_idx += 1
        finally:
            doc.close()

        if not chunks:
            warnings.append("pdf 未提取到任何 chunk")

        return chunks, warnings


class XlsxIngestor:
    """Ingest xlsx sheets into unified rule-learning chunks."""

    def ingest(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"xlsx 文件不存在: {file_path}")

        wb = load_workbook(str(path), data_only=True)
        extracted_at = datetime.utcnow().isoformat()
        chunks: List[Dict[str, Any]] = []
        warnings: List[str] = []
        chunk_idx = 0

        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue

            headers = [safe_str(c.value) for c in ws[1]]
            semantics = [_semantic_of(h) if h else "raw" for h in headers]

            for row_idx in range(2, ws.max_row + 1):
                values = [ws.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
                if not any(v not in (None, "") for v in values):
                    continue

                mapped: Dict[str, str] = {}
                raw_map: Dict[str, str] = {}
                for i, value in enumerate(values):
                    val = safe_str(value)
                    if not val:
                        continue
                    header = headers[i] if i < len(headers) else f"col_{i + 1}"
                    sem = semantics[i] if i < len(semantics) else "raw"
                    raw_map[header or f"col_{i + 1}"] = val
                    if sem != "raw" and sem not in mapped:
                        mapped[sem] = val

                if not raw_map:
                    continue

                key_pairs = [f"{k}: {v}" for k, v in mapped.items()] or [f"{k}: {v}" for k, v in raw_map.items()]
                raw_text = "\n".join(key_pairs)

                chunks.append(
                    {
                        "chunk_id": f"xlsx::{path.name}::{ws.title}::{row_idx}::{chunk_idx}",
                        "input_type": "xlsx",
                        "file_name": path.name,
                        "sheet_name": ws.title,
                        "page": None,
                        "section_title": ws.title,
                        "raw_text": raw_text,
                        "metadata": {
                            "row_index": row_idx,
                            "semantic_fields": mapped,
                            "raw_fields": raw_map,
                        },
                        "extracted_at": extracted_at,
                    }
                )
                chunk_idx += 1

            if ws.max_row > 1 and not any(c.get("sheet_name") == ws.title for c in chunks):
                warnings.append(f"sheet={ws.title}: 未提取到有效数据行")

        if not chunks:
            warnings.append("xlsx 未提取到任何 chunk")

        return chunks, warnings


class FileRouter:
    """Unified ingest entry for xlsx/pdf/txt/md/raw text."""

    def __init__(self):
        self.xlsx_ingestor = XlsxIngestor()
        self.pdf_ingestor = PdfIngestor()
        self.text_ingestor = TextIngestor()

    def ingest(self, source: str, source_name: Optional[str] = None, hints: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        hints = hints or {}
        source = source or ""

        # Case 1: treat source as existing file path.
        path = Path(source)
        if path.exists() and path.is_file():
            suffix = path.suffix.lower()
            if suffix == ".xlsx":
                chunks, warnings = self.xlsx_ingestor.ingest(str(path))
                return {
                    "success": True,
                    "input_type": "xlsx",
                    "file_name": path.name,
                    "chunks": chunks,
                    "chunks_count": len(chunks),
                    "warnings": warnings,
                }
            if suffix == ".pdf":
                chunks, warnings = self.pdf_ingestor.ingest(str(path))
                return {
                    "success": True,
                    "input_type": "pdf",
                    "file_name": path.name,
                    "chunks": chunks,
                    "chunks_count": len(chunks),
                    "warnings": warnings,
                }
            if suffix in {".txt", ".md"}:
                chunks, warnings = self.text_ingestor.ingest_file(str(path))
                return {
                    "success": True,
                    "input_type": "md" if suffix == ".md" else "text",
                    "file_name": path.name,
                    "chunks": chunks,
                    "chunks_count": len(chunks),
                    "warnings": warnings,
                }
            return {
                "success": False,
                "input_type": "unknown",
                "file_name": path.name,
                "chunks": [],
                "chunks_count": 0,
                "warnings": [f"不支持的文件类型: {suffix}"],
            }

        # Case 2: raw text content.
        text = source
        inferred_name = source_name or hints.get("source_name") or "inline_text"
        chunks, warnings = self.text_ingestor.ingest(text=text, file_name=inferred_name, input_type="text")
        return {
            "success": True,
            "input_type": "text",
            "file_name": inferred_name,
            "chunks": chunks,
            "chunks_count": len(chunks),
            "warnings": warnings,
        }
